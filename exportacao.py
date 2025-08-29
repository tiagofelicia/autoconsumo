# exportacao.py

import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime

# ##########################################################################
# FUNÇÃO 1: PARA O SIMULADOR DE TARIFÁRIOS
# ##########################################################################

def criar_excel_para_simulador_tarifarios(df_original, df_simulado, nome_cenario="Cenário Simulado"):
    """
    Gera um ficheiro Excel otimizado para importação no Simulador de Tarifários de Tiago Felícia.
    """
    output = io.BytesIO()

    # Usar o pandas ExcelWriter com o motor openpyxl para ter acesso a formatação
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # --- FOLHA 1: PARA IMPORTAR NO SIMULADOR DE TARIFÁRIOS ---
        df_importavel = pd.DataFrame({
            'Data': df_simulado['DataHora'].dt.strftime('%Y-%m-%d'),
            'Hora': df_simulado['DataHora'].dt.strftime('%H:%M'),
            'Consumo registado, Ativa (kW)': df_original['Consumo_Total_Casa_kWh']*4,
            'Consumo Simulado (kW)': df_simulado['Consumo_Rede_Final_kWh']*4,
        })
        df_importavel.to_excel(writer, sheet_name='Para Importar', index=False, startrow=5)

        # --- FOLHA 2: INSTRUÇÕES ---
        texto_instrucoes = [
            ["Guia de Utilização deste Ficheiro"],
            [],
            ["Folha 'Para Importar'"],
            ["- Esta folha está formatada para ser importada diretamente no Tiago Felícia - Simulador de Tarifários de Eletricidade."],
            ["- A coluna 'Consumo Simulado (kW)' contém os valores do consumo líquido da rede APÓS a simulação."],
            ["- Valores em kW"],
            ["- Use esta folha como se fosse um novo ficheiro de diagrama de carga da E-Redes."],
            [],
            [f"Exportado em: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
        ]
        df_instrucoes = pd.DataFrame(texto_instrucoes)
        df_instrucoes.to_excel(writer, sheet_name='Instruções', index=False, header=False)

        # --- FORMATAÇÃO PROFISSIONAL COM OPENPYXL ---
        workbook = writer.book

        # Formatar a folha 'Para Importar' para se parecer com um ficheiro da E-Redes
        ws_import = writer.sheets['Para Importar']
        ws_import['A1'] = "Relatório de Leituras"
        ws_import['A1'].font = Font(bold=True, size=14)
        ws_import['A3'] = "Diagrama de Carga (Consumo)"
        ws_import['A3'].font = Font(bold=True)
        ws_import['A4'] = f"Cenário: {nome_cenario}"
        
        # Auto-ajustar a largura das colunas em todas as folhas
        for sheet_name in workbook.sheetnames:
            worksheet = writer.sheets[sheet_name]
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

    return output.getvalue()

# ##########################################################################
# FUNÇÃO 2: PARA A ANÁLISE VENDA EXCEDENTE
# ##########################################################################
def criar_excel_analise_venda_excedente(
    df_original, df_simulado, df_omie,
    modelo_venda, tipo_comissao, valor_comissao,
    nome_cenario="Cenário Simulado"
):
    """
    Gera um ficheiro Excel focado na análise financeira do excedente de venda
    e na comparação de dados energéticos, com um sumário dinâmico.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # --- FOLHA 1: ANÁLISE DE VENDA OMIE ---
        df_venda = pd.merge(df_simulado[['DataHora', 'Injecao_Rede_Final_kWh']], df_omie[['DataHora', 'OMIE']], on='DataHora', how='left')
        df_venda = df_venda[df_venda['Injecao_Rede_Final_kWh'] > 0.001].copy()

        df_venda['Preço OMIE (€/MWh)'] = df_venda['OMIE']
        df_venda['Preço OMIE (€/kWh)'] = df_venda['OMIE'] / 1000
        # O nome desta coluna é sempre baseado no OMIE, independentemente do modelo de venda
        df_venda['Valor Bruto Venda OMIE (€)'] = df_venda['Injecao_Rede_Final_kWh'] * df_venda['Preço OMIE (€/kWh)']

        if modelo_venda == 'Preço Fixo':
            df_venda['Valor Líquido Venda (€)'] = df_venda['Injecao_Rede_Final_kWh'] * valor_comissao
        else: # Indexado ao OMIE
            if tipo_comissao == 'Percentual (%)':
                comissao_decimal = valor_comissao / 100.0
                # O cálculo do líquido baseia-se no bruto do OMIE
                df_venda['Valor Líquido Venda (€)'] = df_venda['Valor Bruto Venda OMIE (€)'] * (1 - comissao_decimal)
            else: # Fixo (€/MWh)
                df_venda['Valor Líquido Venda (€)'] = df_venda['Injecao_Rede_Final_kWh'] * ((df_venda['Preço OMIE (€/MWh)'] - valor_comissao) / 1000)
        
        df_venda['Valor Líquido Venda (€)'] = df_venda['Valor Líquido Venda (€)'].clip(lower=0)
        
        # Selecionar e renomear colunas para a exportação
        df_venda_export = df_venda[['DataHora', 'Injecao_Rede_Final_kWh', 'Preço OMIE (€/MWh)', 'Preço OMIE (€/kWh)', 'Valor Bruto Venda OMIE (€)', 'Valor Líquido Venda (€)']].copy()
        df_venda_export.rename(columns={'Injecao_Rede_Final_kWh': 'Injeção Simulada (kWh)'}, inplace=True)
        
        # --- LÓGICA DINÂMICA PARA O SUMÁRIO DE TOTAIS ---
        metricas_totais = []
        valores_totais = []

        # Informação base
        metricas_totais.extend(["Modelo de Venda"])
        valores_totais.extend([modelo_venda])
        soma_injecao = df_venda_export['Injeção Simulada (kWh)'].sum()
        metricas_totais.append('Injeção Total (kWh)')
        valores_totais.append(f"{soma_injecao:,.2f}".replace(",", " ").replace(".", ",") + " kWh")

        # Cálculos totais
        soma_bruto_omie = df_venda_export['Valor Bruto Venda OMIE (€)'].sum()
        soma_liquido = df_venda_export['Valor Líquido Venda (€)'].sum()
        preco_medio_liquido = soma_liquido / soma_injecao if soma_injecao > 0 else 0

        # Bloco condicional para apresentar os totais de forma clara
        if modelo_venda == 'Preço Fixo':
            metricas_totais.append('Preço de Venda Fixo')
            valores_totais.append(f"{valor_comissao:.4f} €/kWh")
            metricas_totais.append("") # Separador
            valores_totais.append("")
            # No preço fixo, o valor líquido é o principal. O OMIE é comparativo.
            metricas_totais.append('Valor Líquido Venda Total (€)')
            valores_totais.append(f"{soma_liquido:,.2f}".replace(",", " ").replace(".", ",") + " €")
            metricas_totais.append('Valor Bruto OMIE Total (€) (p/ comparação)')
            valores_totais.append(f"{soma_bruto_omie:,.2f}".replace(",", " ").replace(".", ",") + " €")
        else: # Indexado ao OMIE
            if tipo_comissao == 'Percentual (%)':
                metricas_totais.append('Comissão Aplicada')
                valores_totais.append(f"{valor_comissao} %")
            else: # Fixo (€/MWh)
                metricas_totais.append('Comissão Aplicada')
                valores_totais.append(f"{valor_comissao:.2f} €/MWh")
            metricas_totais.append("") # Separador
            valores_totais.append("")
            # No modo indexado, o fluxo é Bruto -> Líquido
            metricas_totais.append('Valor Bruto Venda OMIE Total (€)')
            valores_totais.append(f"{soma_bruto_omie:,.2f}".replace(",", " ").replace(".", ",") + " €")
            metricas_totais.append('Valor Líquido Venda Total (€)')
            valores_totais.append(f"{soma_liquido:,.2f}".replace(",", " ").replace(".", ",") + " €")

        # Preço médio líquido é sempre relevante
        metricas_totais.append('Preço Médio Venda Líquido (€/kWh)')
        valores_totais.append(f"{preco_medio_liquido:,.4f}".replace(",", " ").replace(".", ",") + " €")

        totais = {'Métrica': metricas_totais, 'Valor': valores_totais}
        df_totais = pd.DataFrame(totais)
        
        df_venda_export.to_excel(writer, sheet_name='Análise Venda Excedente', index=False)
        df_totais.to_excel(writer, sheet_name='Análise Venda Excedente', index=False, header=False, startrow=len(df_venda_export) + 3)
        
        # --- RESTANTE DO CÓDIGO (FOLHAS E FORMATAÇÃO) ---
        df_comparativo = pd.DataFrame({
            'Data': df_simulado['DataHora'].dt.strftime('%Y-%m-%d'),
            'Hora': df_simulado['DataHora'].dt.strftime('%H:%M'),
            'Consumo registado, Ativa (kWh)': df_original['Consumo_Total_Casa_kWh'],
            'Consumo medido na IC, Ativa (kWh)': df_original['Consumo (kWh)'],
            'Consumo Simulado (kWh)': df_simulado['Consumo_Rede_Final_kWh'],
            'Injeção registada, Ativa (kWh)': df_original['Injecao_Rede_kWh'],
            'Injeção na rede medida na IC, Ativa (kWh)': df_original['Injecao_Total_UPAC_kWh'],
            'Injeção Simulada (kWh)': df_simulado['Injecao_Rede_Final_kWh'],
            'Settlement Original (kWh)': df_original['Consumo_Total_Casa_kWh'] - df_original['Consumo (kWh)'],
            'Settlement Simulado (kWh)': df_original['Consumo_Total_Casa_kWh'] - df_simulado['Consumo_Rede_Final_kWh']
        })
        df_comparativo.to_excel(writer, sheet_name='Dados Comparativos', index=False)
        texto_instrucoes = [
            ["Guia de Utilização deste Ficheiro de Análise de Venda de Excedente"],
            [],
            ["Folha 'Análise Venda Excedente'"],
            ["- Detalhe de cada período de 15 min com injeção de excedente, incluindo preços OMIE e valores de venda brutos e líquidos."],
            ["- Valores em kWh"],
            [],
            ["Folha 'Dados Comparativos'"],
            ["- Contém a comparação detalhada, a cada 15 minutos, entre o consumo original e o consumo após a simulação de autoconsumo."],
            ["- Valores em kWh"],
            [],
            [f"Exportado em: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
        ]
        df_instrucoes = pd.DataFrame(texto_instrucoes)
        df_instrucoes.to_excel(writer, sheet_name='Instruções', index=False, header=False)
        workbook = writer.book
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        for sheet_name in ['Análise Venda Excedente', 'Dados Comparativos']:
            ws = writer.sheets[sheet_name]
            for cell in ws["1:1"]:
                cell.font = header_font
                cell.fill = header_fill
        for sheet_name in workbook.sheetnames:
            worksheet = writer.sheets[sheet_name]
            for col in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

    return output.getvalue()