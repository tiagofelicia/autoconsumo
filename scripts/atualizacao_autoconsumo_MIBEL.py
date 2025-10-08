# --- Carregar as bibliotecas necessárias ---
import pandas as pd
import numpy as np
import requests
import openpyxl
import re

print("✅ Bibliotecas carregadas")

# ===================================================================
# ---- CONFIGURAÇÕES ----
# ===================================================================
DATA_INICIO_ATUALIZACAO = pd.to_datetime("2025-10-01")
FICHEIRO_EXCEL = "Simulador_☀️_Autoconsumo_Tiago_Felicia.xlsx"
ABA_EXCEL = "OMIE_CICLOS"
COLUNA_PARA_ESCREVER = 8 # Coluna H

print(f"ℹ️ Data de início da atualização definida para: {DATA_INICIO_ATUALIZACAO.date()}")
print("⚠️ MODO: Apenas dados reais OMIE (sem futuros)")
# ===================================================================

def run_update_process():
    """
    Função principal que encapsula todo o processo de ETL (SEM FUTUROS).
    """
    try:
        # --- Passo 1: Leitura e Combinação dos Dados OMIE (Tudo em Hora de Espanha) ---
        print("\n⏳ Passo 1: A ler e combinar dados OMIE (todos quarto-horários)...")
        
        fontes_qh = []
        
        print("   - 1a: A ler 'MIBEL.xlsx'...")
        try:
            dados_base_qh = pd.read_excel("MIBEL.xlsx", usecols=['Data', 'Hora', 'Preço marginal no sistema português (EUR/MWh)'])
            dados_base_qh = dados_base_qh.rename(columns={'Preço marginal no sistema português (EUR/MWh)': 'Preco'})
            dados_base_qh['Data'] = pd.to_datetime(dados_base_qh['Data'])
            fontes_qh.append(dados_base_qh)
        except Exception as e: 
            print(f"   - Aviso: Não foi possível ler 'MIBEL.xlsx'. {e}")
        
        print("   - 1b: A ler dados recentes (ACUM)...")
        try:
            dados_acum_qh = pd.read_csv(
                "https://www.omie.es/sites/default/files/dados/NUEVA_SECCION/INT_PBC_EV_H_ACUM.TXT", 
                sep=';', 
                skiprows=2, 
                header=0, 
                usecols=[0, 1, 3], 
                decimal=',', 
                encoding='windows-1252'
            )
            dados_acum_qh.columns = ['Data', 'Hora', 'Preco']
            dados_acum_qh['Data'] = pd.to_datetime(dados_acum_qh['Data'], format='%d/%m/%Y', errors='coerce')
            fontes_qh.append(dados_acum_qh.dropna())
        except Exception as e: 
            print(f"   - Aviso: Falha ao ler dados (ACUM). {e}")

        print("   - 1c: A ler dados do dia seguinte (INDICADORES)...")
        try:
            r = requests.get("https://www.omie.es/sites/default/files/dados/diario/INDICADORES.DAT", timeout=10)
            linhas = r.content.decode('utf-8').splitlines()
            data_sessao = pd.to_datetime([l for l in linhas if l.startswith("SESION;")][0].split(';')[1], format='%d/%m/%Y')
            linhas_dados = [l for l in linhas if re.match(r'^H\d{2}Q[1-4];', l)]
            if linhas_dados:
                dados_ind_list = []
                for l in linhas_dados:
                    partes = l.split(';')
                    hora_str = partes[0]
                    hora = (int(hora_str[1:3])-1)*4 + int(hora_str[4:5])
                    preco = float(partes[2].replace(',', '.'))
                    dados_ind_list.append({'Data': data_sessao, 'Hora': hora, 'Preco': preco})
                fontes_qh.append(pd.DataFrame(dados_ind_list))
        except Exception as e: 
            print(f"   - Aviso: Falha ao ler dados (INDICADORES). {e}")
        
        print("   - 1d: A combinar fontes de dados...")
        todos_dados_qh = pd.concat(fontes_qh, ignore_index=True).drop_duplicates(subset=['Data', 'Hora'], keep='last')
        
        dados_para_manter = todos_dados_qh[todos_dados_qh['Data'] < DATA_INICIO_ATUALIZACAO]
        dados_para_atualizar = todos_dados_qh[todos_dados_qh['Data'] >= DATA_INICIO_ATUALIZACAO]

        dados_combinados_qh = pd.concat([dados_para_manter, dados_para_atualizar]).sort_values(['Data', 'Hora']).reset_index(drop=True)
        
        ultima_data_omie = dados_combinados_qh['Data'].max()
        print(f"✅ Todas as fontes de dados OMIE foram combinadas.")
        print(f"   📅 Última data com dados: {ultima_data_omie.date()}")

        # --- Passo 2: Criar estrutura completa e converter para Portugal ---
        print("\n⏳ Passo 2: A criar estrutura completa até 2026...")
        
        def num_quartos_dia(data):
            """Calcula número de quartos horários considerando DST"""
            tz_es = 'Europe/Madrid'
            dt0 = pd.Timestamp(f"{data} 00:00:00", tz=tz_es)
            dt24 = pd.Timestamp(f"{data} 23:59:59", tz=tz_es)
            horas = (dt24 - dt0).total_seconds() / 3600
            return int(round(horas * 4))
        
        # Gerar estrutura completa de datas futuras (sem preços)
        datas_futuras = pd.date_range(start=ultima_data_omie + pd.Timedelta(days=1), end='2025-12-31', freq='D')
        
        futuro_qh = []
        for data in datas_futuras:
            n_quartos = num_quartos_dia(data)
            for hora in range(1, n_quartos + 1):
                futuro_qh.append({'Data': data, 'Hora': hora, 'Preco': np.nan})
        
        futuro_qh = pd.DataFrame(futuro_qh)
        
        # Combinar histórico + estrutura futura (vazia)
        dados_completos_qh = pd.concat([dados_combinados_qh[['Data', 'Hora', 'Preco']], futuro_qh], ignore_index=True)
        dados_completos_qh = dados_completos_qh.sort_values(['Data', 'Hora']).reset_index(drop=True)
        
        print(f"   - Dados reais até: {ultima_data_omie.date()}")
        print(f"   - Estrutura criada até: 2026-12-31")
        print(f"   - Registos com dados reais: {len(dados_combinados_qh)}")
        print(f"   - Registos vazios (futuros): {len(futuro_qh)}")
        
        print("\n⏳ Passo 3: A converter para hora de Portugal...")
        
        # Gerar datetime em hora de Espanha
        def gerar_datetime_es(row):
            """Gera timestamp correto considerando DST"""
            data = row['Data']
            hora = row['Hora']
            inicio_dia = pd.Timestamp(f"{data} 00:00:00", tz='Europe/Madrid')
            return inicio_dia + pd.Timedelta(minutes=15 * (hora - 1))
        
        dados_completos_qh['datetime_es'] = dados_completos_qh.apply(gerar_datetime_es, axis=1)
        dados_completos_qh['datetime_pt'] = dados_completos_qh['datetime_es'].dt.tz_convert('Europe/Lisbon')
        dados_completos_qh['Data_PT'] = dados_completos_qh['datetime_pt'].dt.date
        
        # Renumerar horas em hora de Portugal
        dados_finais_pt = dados_completos_qh.sort_values('datetime_pt').copy()
        dados_finais_pt['Hora_PT'] = dados_finais_pt.groupby('Data_PT').cumcount() + 1
        
        # Selecionar apenas 2025 e 2026
        dados_finais_pt = dados_finais_pt[dados_finais_pt['datetime_pt'].dt.year.isin([2025, 2026])].copy()
        dados_finais_pt = dados_finais_pt[['Data_PT', 'Hora_PT', 'Preco']].rename(
            columns={'Data_PT': 'Data', 'Hora_PT': 'Hora'}
        )
        dados_finais_pt = dados_finais_pt.reset_index(drop=True)
        
        registos_com_dados = dados_finais_pt['Preco'].notna().sum()
        registos_vazios = dados_finais_pt['Preco'].isna().sum()
        
        print(f"✅ {len(dados_finais_pt)} registos totais preparados em hora de Portugal.")
        print(f"   - Com dados reais: {registos_com_dados}")
        print(f"   - Vazios (NaN): {registos_vazios}")
        
        # Validação de quartos
        check_quartos = dados_finais_pt.groupby('Data').size().reset_index(name='n')
        dias_estranhos = check_quartos[~check_quartos['n'].isin([92, 96, 100])]
        
        if not dias_estranhos.empty:
            print("⚠️ Aviso: Dias com número de quartos inesperado:")
            print(dias_estranhos.to_string(index=False))
        else:
            print("✅ Todos os dias têm número de quartos esperado (92, 96 ou 100).")

        # --- Passo 4: Atualização do ficheiro Excel ---
        print(f"\n⏳ Passo 4: A atualizar o ficheiro '{FICHEIRO_EXCEL}'...")
        
        wb = openpyxl.load_workbook(FICHEIRO_EXCEL)
        sheet = wb[ABA_EXCEL]
        
        # Escrever todos os valores (incluindo NaN que o Excel mostra como vazio)
        for i, preco in enumerate(dados_finais_pt['Preco'].tolist()):
            if pd.isna(preco):
                sheet.cell(row=i + 2, column=COLUNA_PARA_ESCREVER, value=None)
            else:
                sheet.cell(row=i + 2, column=COLUNA_PARA_ESCREVER, value=preco)
        
        # Atualizar data de referência na aba 'Constantes'
        sheet_const = wb["Constantes"]
        sheet_const['B42'] = ultima_data_omie.strftime('%d/%m/%Y')
        
        wb.save(FICHEIRO_EXCEL)
        print(f"✅ O ficheiro Excel foi atualizado com sucesso!")
        print(f"   Data_Valores_OMIE = {ultima_data_omie.date()}")
        print(f"   ⚠️ Nota: Apenas dados reais até {ultima_data_omie.date()} foram escritos.")

    except Exception as e:
        import traceback
        print(f"❌ Ocorreu um erro inesperado: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    run_update_process()
